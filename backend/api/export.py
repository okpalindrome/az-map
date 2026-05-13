"""Export API: JSON (full scan) and Excel workbook (multi-sheet)."""
import io
import json
import re
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session

from ..database import get_db
from ..models.db_models import Edge, Finding, Node, RoleAssignment, Scan

router = APIRouter(prefix="/api/export", tags=["export"])


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe_name(text: str) -> str:
    return re.sub(r'[^\w\-]', '_', text or 'azmap')[:60]


def _load_scan_data(scan_id: str, db: Session):
    scan = db.query(Scan).filter(Scan.id == scan_id).first()
    if not scan:
        raise HTTPException(404, "Scan not found")
    findings = db.query(Finding).filter(Finding.scan_id == scan_id).all()
    nodes = db.query(Node).filter(Node.scan_id == scan_id).all()
    edges = db.query(Edge).filter(Edge.scan_id == scan_id).all()
    ras = db.query(RoleAssignment).filter(RoleAssignment.scan_id == scan_id).all()
    return scan, findings, nodes, edges, ras


# ── JSON export ─────────────────────────────────────────────────────────────────

@router.get("/{scan_id}/json")
def export_json(scan_id: str, db: Session = Depends(get_db)):
    scan, findings, nodes, edges, ras = _load_scan_data(scan_id, db)

    payload = {
        "az_map_version": "1.1",
        "exported_at": datetime.utcnow().isoformat(),
        "scan": {
            "id": scan.id,
            "subscription_id": scan.subscription_id,
            "subscription_name": scan.subscription_name,
            "tenant_id": scan.tenant_id,
            "status": scan.status,
            "started_at": scan.started_at.isoformat() if scan.started_at else None,
            "completed_at": scan.completed_at.isoformat() if scan.completed_at else None,
        },
        "summary": {
            "total_nodes": len(nodes),
            "total_edges": len(edges),
            "total_findings": len(findings),
            "critical_findings": sum(1 for f in findings if f.severity == "critical"),
            "high_findings": sum(1 for f in findings if f.severity == "high"),
        },
        "findings": [
            {
                "id": f.id,
                "finding_type": f.finding_type,
                "severity": f.severity,
                "title": f.title,
                "description": f.description,
                "affected_node": f.affected_node_name,
                "affected_node_id": f.affected_node_id,
                "risk_score": f.risk_score,
                "blast_radius": f.blast_radius,
                "why_risky": f.why_risky,
                "remediation": f.remediation,
                "attack_chain": f.attack_chain,
                "tags": f.tags,
            }
            for f in sorted(findings, key=lambda x: x.risk_score, reverse=True)
        ],
        "nodes": [
            {
                "node_id": n.node_id,
                "node_type": n.node_type,
                "name": n.name,
                "display_name": n.display_name,
                "risk_level": n.risk_level,
                "risk_score": n.risk_score,
                "risk_reasons": n.risk_reasons or [],
                "properties": n.properties or {},
            }
            for n in nodes
        ],
        "edges": [
            {
                "source_node_id": e.source_node_id,
                "target_node_id": e.target_node_id,
                "edge_type": e.edge_type,
                "properties": e.properties or {},
            }
            for e in edges
        ],
        "role_assignments": [
            {
                "principal_id": ra.principal_id,
                "principal_name": ra.principal_name,
                "principal_type": ra.principal_type,
                "role_name": ra.role_name,
                "scope": ra.scope,
                "scope_level": ra.scope_level,
            }
            for ra in ras
        ],
    }
    fname = _safe_name(scan.subscription_name or 'azmap')
    return Response(
        content=json.dumps(payload, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}.json"'},
    )


# ── Excel export ────────────────────────────────────────────────────────────────
# /csv kept as the route name for UI compatibility; returns .xlsx

@router.get("/{scan_id}/csv")
def export_excel(scan_id: str, db: Session = Depends(get_db)):
    """Export scan data as a multi-sheet Excel workbook (.xlsx).

    Sheet 1 – Findings   : severity, risk score, title, type, affected resource name,
                           affected resource ID, blast radius, subscription name
    Sheet 2 – Inventory  : node ID, node type, name, display name, risk level,
                           risk score, risk reasons
    Sheet 3 – Roles      : principal name, principal type, principal ID,
                           role name, scope, scope level
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    scan, findings, nodes, _, ras = _load_scan_data(scan_id, db)
    sub_name = scan.subscription_name or scan.subscription_id or ""

    wb = openpyxl.Workbook()

    # ── Styling helpers ────────────────────────────────────────────────────────
    SEV_FILL = {
        "critical": "FFEBEE",
        "high":     "FFF3E0",
        "medium":   "FFFDE7",
        "low":      "E8F5E9",
        "info":     "F5F5F5",
    }
    HEADER_FILL = PatternFill("solid", fgColor="18181B")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

    def _header_row(ws, cols: list[str]):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left")

    def _col_widths(ws, widths: list[int]):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _freeze(ws):
        ws.freeze_panes = "A2"

    # ── Sheet 1: Findings ──────────────────────────────────────────────────────
    ws_f = wb.active
    ws_f.title = "Findings"

    _header_row(ws_f, [
        "Severity", "Risk Score", "Title", "Type",
        "Affected Resource", "Affected Resource ID",
        "Blast Radius", "Subscription",
    ])

    for f in sorted(findings, key=lambda x: x.risk_score, reverse=True):
        row = [
            f.severity or "",
            f.risk_score or 0,
            f.title or "",
            (f.finding_type or "").replace("_", " "),
            f.affected_node_name or "",
            f.affected_node_id or "",
            f.blast_radius or 0,
            sub_name,
        ]
        ws_f.append(row)
        # Colour-code the severity cell
        sev = (f.severity or "").lower()
        if sev in SEV_FILL:
            ws_f.cell(ws_f.max_row, 1).fill = PatternFill("solid", fgColor=SEV_FILL[sev])

    _col_widths(ws_f, [12, 10, 48, 28, 32, 40, 12, 28])
    _freeze(ws_f)

    # ── Sheet 2: Inventory ────────────────────────────────────────────────────
    ws_i = wb.create_sheet("Inventory")

    _header_row(ws_i, [
        "Node ID", "Node Type", "Name", "Display Name",
        "Risk Level", "Risk Score", "Risk Reasons",
    ])

    for n in sorted(nodes, key=lambda x: x.risk_score, reverse=True):
        ws_i.append([
            n.node_id or "",
            (n.node_type or "").replace("_", " "),
            n.name or "",
            n.display_name or "",
            n.risk_level or "",
            n.risk_score or 0,
            "; ".join(n.risk_reasons or []),
        ])

    _col_widths(ws_i, [40, 18, 30, 30, 10, 10, 60])
    _freeze(ws_i)

    # ── Sheet 3: Roles ────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("Roles")

    _header_row(ws_r, [
        "Principal Name", "Principal Type", "Principal ID",
        "Role Name", "Scope", "Scope Level",
    ])

    for ra in sorted(ras, key=lambda x: x.role_name or ""):
        ws_r.append([
            ra.principal_name or "",
            ra.principal_type or "",
            ra.principal_id or "",
            ra.role_name or "",
            ra.scope or "",
            ra.scope_level or "",
        ])

    _col_widths(ws_r, [32, 18, 40, 36, 70, 14])
    _freeze(ws_r)

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = _safe_name(sub_name)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
    )
